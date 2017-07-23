"""
Some codes from https://github.com/Newmu/dcgan_code
"""
from __future__ import division
import math
import json
import random
import pprint
import scipy.misc
import numpy as np
from time import gmtime, strftime
from six.moves import xrange

import tensorflow as tf
import tensorflow.contrib.slim as slim
import pickle
import xlwt
import os

import matplotlib.gridspec as gridspec
import matplotlib.pyplot as plt
from scipy import stats

from openpyxl import Workbook
from openpyxl import load_workbook



pp = pprint.PrettyPrinter()

get_stddev = lambda x, k_h, k_w: 1/math.sqrt(k_w*k_h*x.get_shape()[-1])

def load_tabular_data(dataset_name, classes):

  with open('./data/'+ dataset_name+ '/'+ dataset_name + '_final.pickle', 'rb') as handle:
            X = pickle.load(handle)
  print( "Data Shape = " + str(X.shape))
  # X should be numpy Array
  with open('./data/'+ dataset_name+ '/'+ dataset_name + '_label.pickle', 'rb') as handle:
            y = pickle.load(handle)


  y = y.reshape(y.shape[0], -1).astype(np.int16)

  print("X1 = " + str( X.shape))
  print("y1 = " + str( y.shape))
  print(classes)

  # with open('data/test_data_' + dataset_name+ '.pickle', 'rb') as handle:
  #          X_test = pickle.load(handle)
  # with open('data/test_label_' + dataset_name+ '.pickle', 'rb') as handle:
  #          y_test = pickle.load(handle)

  
  y_onehot = np.zeros((len(y), classes), dtype=np.float)
  for i, lbl in enumerate(y):
      y_onehot[i, y[i]] = 1.0

  print( "y Shape = " + str(y_onehot.shape))

  return X , y_onehot, y
      
def show_all_variables():
  model_vars = tf.trainable_variables()
  slim.model_analyzer.analyze_vars(model_vars, print_info=True)

def get_image(image_path, input_height, input_width,
              resize_height=64, resize_width=64,
              crop=True, grayscale=False):
  image = imread(image_path, grayscale)
  return transform(image, input_height, input_width,
                   resize_height, resize_width, crop)

def save_images(images, size, image_path):
  return imsave(inverse_transform(images), size, image_path)

def save_data(data, data_file):
  with open(data_file, 'wb') as handle:
    return pickle.dump(data, handle, protocol=pickle.HIGHEST_PROTOCOL)
  

def imread(path, grayscale = False):
  if (grayscale):
    return scipy.misc.imread(path, flatten = True).astype(np.float)
  else:
    return scipy.misc.imread(path).astype(np.float)

def merge_images(images, size):
  return inverse_transform(images)

def merge(images, size):
  h, w = images.shape[1], images.shape[2]
  if (images.shape[3] in (3,4)):
    c = images.shape[3]
    img = np.zeros((h * size[0], w * size[1], c))
    for idx, image in enumerate(images):
      i = idx % size[1]
      j = idx // size[1]
      img[j * h:j * h + h, i * w:i * w + w, :] = image
    return img
  elif images.shape[3]==1:
    img = np.zeros((h * size[0], w * size[1]))
    for idx, image in enumerate(images):
      i = idx % size[1]
      j = idx // size[1]
      img[j * h:j * h + h, i * w:i * w + w] = image[:,:,0]
    return img
  else:
    raise ValueError('in merge(images,size) images parameter '
                     'must have dimensions: HxW or HxWx3 or HxWx4')

def imsave(images, size, path):
  image = np.squeeze(merge(images, size))
  return scipy.misc.imsave(path, image)

def center_crop(x, crop_h, crop_w,
                resize_h=64, resize_w=64):
  if crop_w is None:
    crop_w = crop_h
  h, w = x.shape[:2]
  j = int(round((h - crop_h)/2.))
  i = int(round((w - crop_w)/2.))
  return scipy.misc.imresize(
      x[j:j+crop_h, i:i+crop_w], [resize_h, resize_w])

def transform(image, input_height, input_width, 
              resize_height=64, resize_width=64, crop=True):
  if crop:
    cropped_image = center_crop(
      image, input_height, input_width, 
      resize_height, resize_width)
  else:
    cropped_image = scipy.misc.imresize(image, [resize_height, resize_width])
  return np.array(cropped_image)/127.5 - 1.

def inverse_transform(images):
  return (images+1.)/2.

def to_json(output_path, *layers):
  with open(output_path, "w") as layer_f:
    lines = ""
    for w, b, bn in layers:
      layer_idx = w.name.split('/')[0].split('h')[1]

      B = b.eval()

      if "lin/" in w.name:
        W = w.eval()
        depth = W.shape[1]
      else:
        W = np.rollaxis(w.eval(), 2, 0)
        depth = W.shape[0]

      biases = {"sy": 1, "sx": 1, "depth": depth, "w": ['%.2f' % elem for elem in list(B)]}
      if bn != None:
        gamma = bn.gamma.eval()
        beta = bn.beta.eval()

        gamma = {"sy": 1, "sx": 1, "depth": depth, "w": ['%.2f' % elem for elem in list(gamma)]}
        beta = {"sy": 1, "sx": 1, "depth": depth, "w": ['%.2f' % elem for elem in list(beta)]}
      else:
        gamma = {"sy": 1, "sx": 1, "depth": 0, "w": []}
        beta = {"sy": 1, "sx": 1, "depth": 0, "w": []}

      if "lin/" in w.name:
        fs = []
        for w in W.T:
          fs.append({"sy": 1, "sx": 1, "depth": W.shape[0], "w": ['%.2f' % elem for elem in list(w)]})

        lines += """
          var layer_%s = {
            "layer_type": "fc", 
            "sy": 1, "sx": 1, 
            "out_sx": 1, "out_sy": 1,
            "stride": 1, "pad": 0,
            "out_depth": %s, "in_depth": %s,
            "biases": %s,
            "gamma": %s,
            "beta": %s,
            "filters": %s
          };""" % (layer_idx.split('_')[0], W.shape[1], W.shape[0], biases, gamma, beta, fs)
      else:
        fs = []
        for w_ in W:
          fs.append({"sy": 5, "sx": 5, "depth": W.shape[3], "w": ['%.2f' % elem for elem in list(w_.flatten())]})

        lines += """
          var layer_%s = {
            "layer_type": "deconv", 
            "sy": 5, "sx": 5,
            "out_sx": %s, "out_sy": %s,
            "stride": 2, "pad": 1,
            "out_depth": %s, "in_depth": %s,
            "biases": %s,
            "gamma": %s,
            "beta": %s,
            "filters": %s
          };""" % (layer_idx, 2**(int(layer_idx)+2), 2**(int(layer_idx)+2),
               W.shape[0], W.shape[3], biases, gamma, beta, fs)
    layer_f.write(" ".join(lines.replace("'","").split()))

def make_gif(images, fname, duration=2, true_image=False):
  import moviepy.editor as mpy

  def make_frame(t):
    try:
      x = images[int(len(images)/duration*t)]
    except:
      x = images[-1]

    if true_image:
      return x.astype(np.uint8)
    else:
      return ((x+1)/2*255).astype(np.uint8)

  clip = mpy.VideoClip(make_frame, duration=duration)
  clip.write_gif(fname, fps = len(images) / duration)

def compare(max_col, config , real , fake , save_dir):

  #gs = gridspec.GridSpec(max_col, 2)
    
  if os.path.isfile("./Evaluations/results.xlsx"):
    wb = load_workbook("./Evaluations/results.xlsx") 
  else:
    wb = Workbook() 

  sh = wb.active

  xrow= sh.max_row 
  # sh = book.add_sheet("results")
# sh.cell
# c =sh.cell( row= xrow , column=1) 
# c.value= "test_id" + str(xrow)

# c =sh.cell( row= xrow + 1, column=1) 
# c.value= "test_id X " + str(xrow + 1 )

# wb.save("resultMM.xlsx")

  # if os.path.isfile(save_dir + "/result.xlsx"):
  #   wb = load_workbook(save_dir + "/result.xlsx") 
  # else:
  #   wb = Workbook() 

  # sh = wb.active

  # sh = book.add_sheet("results")

  sh.cell( row = 1  , column =1).value = "test_id"
  sh.cell( row = xrow+1  , column =1).value =config.test_id
  sh.cell( row = 1  , column =2).value = "alpha"
  sh.cell( row = xrow+1  , column =2).value = config.alpha
  sh.cell( row = 1  , column =3).value = "beta"
  sh.cell( row = xrow+1  , column =3).value = config.beta
  sh.cell( row = 1  , column =4).value = "delat_mean"
  sh.cell( row = xrow +1 , column =4).value = config.delta_m
  sh.cell( row = 1  , column =5).value = "delta_var"
  sh.cell( row = xrow+1  , column =5).value = config.delta_v

    # sh.write( 1 , 0, config.test_id)
  # sh.write( 0 , 1, "alpha")
  # sh.write( 1 , 1, config.alpha)
  # sh.write( 0 , 2, "beta")
  # sh.write( 1 , 2, config.beta)
  # sh.write( 0 , 3, "delta_mean")
  # sh.write( 1 , 3, config.delta_m)
  # sh.write( 0 , 4, "delta_var")
  # sh.write( 1 , 4, config.delta_v)
  
   
  if not os.path.exists(save_dir+'/histo'):
          os.makedirs(save_dir+'/histo')  

  for i in range(max_col):   
    
      #plt.subplots_adjust(top=3.92, bottom=0.08, left=0.10, right=0.95, hspace=0.25,
      #                wspace=0.25)
      
    
      
      plt.figure() # <- makes new figure and makes it active (remove this)
      plt.hist( real[:,i] , bins='auto')
      
      plt.xlabel('Real Values. Column = ' + str(i + 1) )
      plt.ylabel('(1,-1) range')
      
      plt.savefig(save_dir + "/histo/"+ str(i + 1) + "_R" ) # <- saves the currently active figure (which is empty in your code)

      # ax =plt.subplot(gs[i,0] )
      
      # ax.set_title("Original: Col " + str( i +1))
      # plt.hist( real[:,i] , bins='auto')

      
      # fig1 = plt.gcf()
      
      # fig1.savefig(save_dir + "/histo/R_"+ str(i + 1) , dpi= 200)
      
      
      # ax =plt.subplot(gs[i,1] )
    
      plt.figure() # <- makes new figure and makes it active (remove this)
      plt.hist( fake[:,i] , bins='auto')
      
      plt.xlabel('Fake Values. Column = '+ str(i + 1))
      plt.ylabel('(1,-1) range')
      
      plt.savefig(save_dir + "/histo/"+ str(i + 1)+"_F" ) 

      print(" Histogram of col " + str(i + 1))
      # ax.set_title("Generated: Col " + str(i+1))
      # plt.hist( fake[:,i] , bins='auto')
      
      # fig1 = plt.gcf()
      # fig1.savefig(save_dir + "/histo/F_"+ str(i + 1 ), dpi= 200)
      
      

      #plt.show()
      
      #s , p =stats.ttest_ind(real[:,cols[i]], fake[:,cols[i]])  
      #print ( "t-test -> t-Statistic = %1.5f , P_Value =%.5f " % (s , p))
      
      s , p =stats.ttest_ind(real[:,i], fake[:,i] , equal_var = False)  

      sh.cell( row = 1  , column =(i*3) + 6).value = "pvalue_" +str( i + 1)
      sh.cell( row = xrow+1  , column =(i*3) + 6).value = float(p)


      # sh.write( 0 , (i*3) + 5, "pvalue_" +str( i + 1))
      # sh.write( 1 , (i*3)+ 5, float(p) )

      mean_col = np.mean(fake[:,i])

      sh.cell( row = 1  , column =(i*3) + 7).value = "mean_fake: "+str( i + 1)
      sh.cell( row = xrow+1  , column =(i*3) + 7).value = float(mean_col)

      # sh.write( 0 , (i*3) + 6, "mean_fakel: "+str( i + 1))
      # sh.write( 1 , (i*3) + 6, float(mean_col) )

      mean_col = np.mean(real[:,i])

      sh.cell( row = 1  , column =(i*3) + 8).value = "mean_real: "+str( i + 1)
      sh.cell( row = xrow+1  , column =(i*3) + 8).value = float(mean_col)

      # sh.write( 0 , (i*3) + 7, "mean_real: "+str( i + 1))
      # sh.write( 1 , (i*3) + 7, float(mean_col) )


      #print ( "Welch -> t-Statistic = %1.5f , P_Value =%.5f " % (s , p))
  wb.save("./Evaluations/results.xlsx")
  print( "Results.xlsx was saved.")
    
def visualize(sess, dcgan, config, option):

  image_frame_dim = int(math.ceil(config.batch_size**.5))

  if option == 0:
    z_sample = np.random.uniform(-0.5, 0.5, size=(config.batch_size, dcgan.z_dim))
    samples = sess.run(dcgan.sampler, feed_dict={dcgan.z: z_sample})
    save_images(samples, [image_frame_dim, image_frame_dim], './samples/test_%s.png' % strftime("%Y%m%d%H%M%S", gmtime()))

  elif option == 1:

    if config.dataset in [ 'mnist', 'celebA' ]: 

      values = np.arange(0, 1, 1./config.batch_size)
      for idx in xrange(100):
        print(" [*] %d" % idx)
        z_sample = np.zeros([config.batch_size, dcgan.z_dim])
        for kdx, z in enumerate(z_sample):
          z[idx] = values[kdx]

        if config.dataset == "mnist":
          y = np.random.choice(10, config.batch_size)
          y_one_hot = np.zeros((config.batch_size, 10))
          y_one_hot[np.arange(config.batch_size), y] = 1

          samples = sess.run(dcgan.sampler, feed_dict={dcgan.z: z_sample, dcgan.y: y_one_hot})
        else:
          samples = sess.run(dcgan.sampler, feed_dict={dcgan.z: z_sample})

        save_images(samples, [image_frame_dim, image_frame_dim], './samples/test_arange_%s.png' % (idx))

      
    else:
      input_size = 15000;  
      dim= config.output_width #16

      merged_data =np.ndarray([config.batch_size * (input_size // config.batch_size) , dim ,dim], dtype=float) # 64 * 234, 16 , 16
      
      save_dir = './{}'.format(config.sample_dir + "/" +config.test_id)

      if not os.path.exists(save_dir):
            os.makedirs(save_dir)  

      samples_dir = save_dir + '/samples'
      if not os.path.exists(samples_dir):
            os.makedirs(samples_dir)  

      for idx in xrange( input_size //config.batch_size) :

        print(" [*] %d" % idx)
        z_sample = np.random.uniform(-1, 1, size=(config.batch_size, dcgan.z_dim))

        if config.dataset == "LACity":
          # Please change this random.choice in a way that the ratio of 0 (poor) to 1 (rich) is the same as the distribution in the original table
          # You don't need to use the label in the original table. Once the ratio of 0 to 1 is the same, there is no problem.
          #y = np.random.choice(2, config.batch_size)
          # 0s = 7164 / 15000 = 0.4776 %, 1s= 7836 / 15000 of data. Similar to oroginal labels
          zero_labeles= 0.48
          y = np.ones((config.batch_size,1))
          
          y[: int(zero_labeles * config.batch_size)] = 0
          np.random.shuffle(y)

          print( "y shape " + str( y.shape))
          y=y.astype('int16')

          #y = np.random.choice(2, 64 , p =[ 0.48 , 1 - 0.48] ) # Not Accurate 
          y_one_hot = np.zeros((config.batch_size, dcgan.y_dim))

          y_one_hot[np.arange(config.batch_size), y] = 1

          samples = sess.run(dcgan.sampler, feed_dict={dcgan.z: z_sample, dcgan.y: y_one_hot , dcgan.y_normal : y })

        else:
          samples = sess.run(dcgan.sampler, feed_dict={dcgan.z: z_sample})
        

        save_data(samples,  samples_dir + '/sample_{:04d}.pickle'.format(idx))

        # Merging Data 
        merged_data[idx * config.batch_size : (idx+1) * config.batch_size] = samples.reshape(samples.shape[0],samples.shape[1],samples.shape[2]) # 64 ,16,16

        print(" == Samples Saved = %4d " % samples.shape[0])
       
      # merged_data is ready in merged_data , now reshape it to a tabular data
      print( "Merged Data Shape = " + str(merged_data.shape))
      tabular_fake_data = merged_data.reshape( merged_data.shape[0] , merged_data.shape[1] * merged_data.shape[2]) # [64 , 16*16=256]

      print( " Fake Data Shape= " + str(tabular_fake_data.shape))

      with open(save_dir +'/fake_tabular.pickle', 'wb') as handle:
        pickle.dump(tabular_fake_data , handle, protocol=pickle.HIGHEST_PROTOCOL)

      real_X, real_y , real_y_normal = load_tabular_data(config.dataset , dcgan.y_dim)

      tabular_real_data = real_X.reshape( real_X.shape[0] , real_X.shape[1] * real_X.shape[2]) # [64 , 16*16=256]

      with open(save_dir +'/real_tabular.pickle', 'wb') as handle:       
        pickle.dump(tabular_real_data , handle, protocol=pickle.HIGHEST_PROTOCOL)

      
      print( "Real Data Shape = " + str(real_X.shape))

      compare(config.maxcol , config , tabular_real_data , tabular_fake_data , save_dir)

    #  for table in samples:
    # values = np.arange(0, 1, 1./config.batch_size)
    # for idx in xrange(100):
    #   print(" [*] %d" % idx)
    #   z_sample = np.zeros([config.batch_size, dcgan.z_dim])
    #   for kdx, z in enumerate(z_sample):
    #     z[idx] = values[kdx]

    #   if config.dataset == "mnist":
    #     y = np.random.choice(10, config.batch_size)
    #     y_one_hot = np.zeros((config.batch_size, 10))
    #     y_one_hot[np.arange(config.batch_size), y] = 1

    #     samples = sess.run(dcgan.sampler, feed_dict={dcgan.z: z_sample, dcgan.y: y_one_hot})
    #   else:
    #     samples = sess.run(dcgan.sampler, feed_dict={dcgan.z: z_sample})

    #   save_images(samples, [image_frame_dim, image_frame_dim], './samples/test_arange_%s.png' % (idx))
  elif option == 2:
    values = np.arange(0, 1, 1./config.batch_size)
    for idx in [random.randint(0, 99) for _ in xrange(100)]:
      print(" [*] %d" % idx)
      z = np.random.uniform(-0.2, 0.2, size=(dcgan.z_dim))
      z_sample = np.tile(z, (config.batch_size, 1))
      #z_sample = np.zeros([config.batch_size, dcgan.z_dim])
      for kdx, z in enumerate(z_sample):
        z[idx] = values[kdx]

      if config.dataset == "mnist":
        y = np.random.choice(10, config.batch_size)
        y_one_hot = np.zeros((config.batch_size, 10))
        y_one_hot[np.arange(config.batch_size), y] = 1

        samples = sess.run(dcgan.sampler, feed_dict={dcgan.z: z_sample, dcgan.y: y_one_hot})
      else:
        samples = sess.run(dcgan.sampler, feed_dict={dcgan.z: z_sample})

      try:
        make_gif(samples, './samples/test_gif_%s.gif' % (idx))
      except:
        save_images(samples, [image_frame_dim, image_frame_dim], './samples/test_%s.png' % strftime("%Y%m%d%H%M%S", gmtime()))
  elif option == 3:
    values = np.arange(0, 1, 1./config.batch_size)
    for idx in xrange(100):
      print(" [*] %d" % idx)
      z_sample = np.zeros([config.batch_size, dcgan.z_dim])
      for kdx, z in enumerate(z_sample):
        z[idx] = values[kdx]

      samples = sess.run(dcgan.sampler, feed_dict={dcgan.z: z_sample})
      make_gif(samples, './samples/test_gif_%s.gif' % (idx))
  elif option == 4:
    image_set = []
    values = np.arange(0, 1, 1./config.batch_size)

    for idx in xrange(100):
      print(" [*] %d" % idx)
      z_sample = np.zeros([config.batch_size, dcgan.z_dim])
      for kdx, z in enumerate(z_sample): z[idx] = values[kdx]

      image_set.append(sess.run(dcgan.sampler, feed_dict={dcgan.z: z_sample}))
      make_gif(image_set[-1], './samples/test_gif_%s.gif' % (idx))

    new_image_set = [merge(np.array([images[idx] for images in image_set]), [10, 10]) \
        for idx in range(64) + range(63, -1, -1)]
    make_gif(new_image_set, './samples/test_gif_merged.gif', duration=8)
